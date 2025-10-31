import pandas as pd
import csv
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
import os
import re
import math
from config.settings import settings

# Define widths based on headers
COLUMN_WIDTHS = {
    "Module Name": 25,
    "Tasks": 20,
    "Subtasks": 20,
    "Deliverables": 25,
    "Epic": 20,
    "ID": 12,
    "User Story": 30,
    "Tasks/Subtasks": 25,
    "Acceptance Criteria": 50,
    "Priority": 15,
    "Story Points": 15,
    "Sprint Number": 15,
    "Assignee": 20,
    # Weeks: keep compact
    "Epic / Week": 20,
    **{f"Week {i}": 12 for i in range(1, 11)}
}

def colorGanttChart(input_file: str, output_file: str = None):
    """
    Colour GanttChart Excel file. 

    Args:
        input_file (str): Path to the input Excel file.
        output_file (str): Path where the Excel file will be saved.
    """
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active  # or wb["Sheet1"] if you know the sheet name

    # Define fills
    header_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")  # yellow
    block_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")  # green

    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column), start=1):
        for cell in row:
            val = str(cell.value) if cell.value is not None else ""

            # Color header row
            if r_idx == 1:
                cell.fill = header_fill

            # Color "█" blocks
            if "█" in val:
                cell.fill = block_fill
                cell.value = ""  # optional: remove the block char, keep just color

    # Save (overwrite or new file)
    if not output_file:
        output_file = input_file
    wb.save(output_file)
    print(f"Coloured GanttChart {input_file} saved at {output_file}")

def addBorder(input_file: str, output_file: str = None):
    """
    Add Border to Excel file. 

    Args:
        input_file (str): Path to the input Excel file.
        output_file (str): Path where the Excel file will be saved.
    """

    wb = openpyxl.load_workbook(input_file)
    ws = wb.active  # or wb["Sheet1"] if you know the sheet name

    # Define border style (thin black lines)
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000")
    )

    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column), start=1):
        for cell in row:
            # add cell boarder
            cell.border = thin_border
            # Color header row
            if r_idx == 1:
                cell.font = Font(bold=True)
     # Save (overwrite or new file)
    if not output_file:
        output_file = input_file
    wb.save(output_file)
    print(f"Added Border to {input_file} saved at {output_file}")

def csv_to_excel(input_file: str, output_file: str):
    """
    Convert a CSV file into an Excel (.xlsx) file.
    
    Args:
        input_file (str): Path to the input CSV file.
        output_file (str): Path where the Excel file will be saved.
    """
    # Create a new workbook
    wb = Workbook()
    ws = wb.active

    # Open the CSV and read line by line
    with open(input_file, "r", newline="", encoding="utf-8") as f:
        reader = csv.reader(f, delimiter="|")

        headers = None  # First row (header row)
        for row in reader:
            if row == ['```csv'] or row ==  ['```']:
                continue
            if headers == None:
                headers = row
            # print(type(row)) //<class 'list'>
            ws.append(row)

    # Assign column widths based on header names
    for idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(idx)
        width = COLUMN_WIDTHS.get(header.strip(), 20)
        ws.column_dimensions[col_letter].width = width
        

    # Apply wrap text + dynamic row height
    base_height = 15  # one line height
    chars_per_line = 40  # rough estimate before wrap
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        total_chars = sum(len(str(cell.value)) for cell in row if cell.value)
        # estimate lines needed
        lines = max(1, math.ceil(total_chars / chars_per_line))
        row_height = base_height * lines
        # set height
        ws.row_dimensions[row[0].row].height = row_height
        # wrap text
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Save as Excel
    wb.save(output_file)
    print(f"Converted {input_file} -> {output_file}")

    #decorate excel file (add border and colors)
    if output_file == "./outputs/GanttChart.xlsx":
        colorGanttChart(output_file)
    addBorder(output_file)
    
def is_file_empty_getsize(file_path):
    """Checks if a file is empty using os.path.getsize()."""
    try:
        return os.path.getsize(file_path) == 0
    except FileNotFoundError:
        print(f"Error: File '{file_path}' not found.")
        return False # Or raise an exception, depending on desired behavior
    except OSError as e:
        print(f"Error accessing file '{file_path}': {e}")
        return False

if __name__ == "__main__":
    csv_to_excel( settings.output_file1, settings.output_file1_xlsx) 
    csv_to_excel( settings.output_file2, settings.output_file2_xlsx) 
    csv_to_excel( settings.output_file3, settings.output_file3_xlsx) 

    # colorGanttChart("./outputs/GanttChart.xlsx", "./outputs/Coloured-GanttChart.xlsx")
    # addBorder("./outputs/UserStories.xlsx", "./outputs/Border.xlsx")